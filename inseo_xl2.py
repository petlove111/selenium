
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serv=Service()
serv.creation_flags=CREATE_NO_WINDOW
driver=webdriver.Chrome(service=serv)

import time,datetime,os
import openpyxl

now = str(datetime.datetime.now())[:16]
folderName = now.replace(":", "_")
os.mkdir(folderName)


now = str(datetime.datetime.now())[:16]
folderName = now.replace(":", "_")
os.mkdir(folderName)


key_word = ["6차 산업"]

wb = openpyxl.Workbook()  

for i in range(len(key_word)):
    ws = wb.create_sheet()
    ws.title = key_word[i]
    ws.column_dimensions["A"].width = 90

    url="https://search.naver.com/search.naver?where=news&sm=tab_jum&query="+key_word[i]
    driver.get(url)
    time.sleep(2)

    list_news = driver.find_element("class name", "list_news")
    news_boxes = list_news.find_elements("class name", "bx")

    for j in range(len(news_boxes)):
        driver.execute_script("arguments[0].scrollIntoView(true);", news_boxes[j])
        file = f"{folderName}/{i+1}_{key_word[i]}-{j+1}.png"
        news_boxes[j].screenshot(file)

        ws.row_dimensions[j+1].height=100
        img = openpyxl.drawing.image.Image(file)
        ws.add_image(img, f'A{j+1}')
        

        title = news_boxes[j].find_element("class name", "news_tit") 
        print(j+1, title.text)
        
        link = title.get_attribute("href")
        ws[f'B{j+1}'].value = "기사링크"   
        ws[f'B{j+1}'].hyperlink = link


    print()

wb.remove(wb["Sheet"])
wb.save(f"{forderName}/{folderName}_결과.xlsx")
