
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serv=Service()
serv.creation_flags=CREATE_NO_WINDOW
driver=webdriver.Chrome(service=serv)



import time, datetime, os #time, datetime, os 불러오기
import openpyxl   


now = str(datetime.datetime.now())[:16] #시간을 15번째 자리까지 now에 저장
folderName = now.replace(":", "_") #:을 _로 바꾸기
os.mkdir(folderName) #폴더 만들기


key_word = ["6차 산업"]

wb = openpyxl.Workbook() #워크북

for i in range(len(key_word)): #키워드 수만큼 반복
    ws = wb.create_sheet() #워크 시트
    ws.title = key_word[i] #워크 시트 이름을 키워드로 설정
    ws.column_dimensions["A"].width = 90 #A행의 너비 조절
    
    url = "https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=" + key_word[i] #검색어를 검색한 페이지의 url
    driver.get(url)
    time.sleep(2)
                  
    list_news = driver.find_element("class name", "list_news") #class name에서 list_news를 list_news에 저장
    news_boxes = list_news.find_elements("class name", "bx") #class name에서 bx를 news_boxes에 저장
   
    for j in range(len(news_boxes)): #중첩 for문
        driver.execute_script("arguments[0].scrollIntoView(true);", news_boxes[j]) #스크롤 내리기
        file = f"{folderName}/{i+1}_{key_word[i]}-{j+1}.png"  #파일 이름 미리 설정
        news_boxes[j].screenshot(file)  #스크린샷

        ws.row_dimensions[j+1].height = 100  #높이 설정
        img = openpyxl.drawing.image.Image(file) #엑셀에서 사용가능한 이미지로 전환
        ws.add_image(img, f'A{j+1}') #A열에 이미지 삽입


        title = news_boxes[j].find_element("class name", "news_tit") #class name에서 news_tit를 title에저장
        print(j+1, title.text) #shell 모드에 출력
        
        link = title.get_attribute("href") #하이퍼링크를 link에 저장
        ws[f'B{j+1}'].value = "기사링크" #B열에  "기사링크"
        ws[f'B{j+1}'].hyperlink = link #하이퍼링크


    print() #줄 띄우기
    
wb.remove(wb["Sheet"]) #필요없는 시트 삭제
wb.save(f"{folderName}/{folderName}_결과.xlsx") #엑셀 파일 저장
